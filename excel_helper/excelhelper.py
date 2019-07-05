#-*- coding: utf-8 -*
import xlrd
import openpyxl


class ExcelReader:
    def __init__(self, filename, encoding=None):
        if not encoding is None:
            xlrd.Book.encoding = 'gbk'
        self.book = xlrd.open_workbook(filename, encoding_override=encoding)
        self.current_sheet = self.book.sheet_by_index(0)

    def set_current_sheet(self, index=None, name=None):
        if index is None and name is None:
            raise ValueError('必须传一个值')
        elif name is None:
            self.current_sheet = self.book.sheet_by_index(index)
        elif index is None:
            self.current_sheet = self.book.sheet_by_name(name)

    def rows(self):
        rows_num = self.current_sheet.nrows
        for i in range(1, rows_num):
            yield self.current_sheet.row_values(i)

    def dict_rows(self):
        for t in self.rows():
            yield dict(zip(self.headers, t))

    @property
    def headers(self):
        if self.row_num == 0:
            return []
        return self.current_sheet.row_values(0)

    @property
    def row_num(self):
        return  self.current_sheet.nrows

    @property
    def col_num(self):
        return self.current_sheet.ncols

    def cell_value(self, row, col):
        return self.current_sheet.cell_value(row, col)

class ExcelWriter:
    def __init__(self, filename, headers=None):
        self.workbook = openpyxl.Workbook()
        self.sheet_num = 0
        self.headers = headers

        if filename.endswith('.xlsx'):
            self.filename = filename
        else:
            self.filename = filename+'.xlsx'
        self.current_row = 1

    def add_sheet(self, name, headers=None):
        if headers is None:
            headers = self.headers
        else:
            self.headers = headers

        self.sheet = self.workbook.create_sheet(name, self.sheet_num)
        if not (self.headers is None and headers is None):
            for i,t in enumerate(headers):
                self.sheet.cell(1, i+1, t)
            self.current_row += 1
            self.sheet_num += 1

    def save(self):
        self.workbook.save(self.filename)

    def write(self, data):
        if not hasattr(self, 'headers'):
            raise ValueError('please set headers before write data into excel.')
        if not hasattr(self, 'sheet'):
            self.add_sheet('Sheet1')

        if not self.headers is None and len(data) != len(self.headers):
            raise ValueError('长度不匹配')
        if isinstance(data, dict):
            for i,t in enumerate(self.headers):
                if data[t] == None:
                    self.sheet.cell(self.current_row, i+1, '')
                else:
                    self.sheet.cell(self.current_row, i+1, data[t])
        elif isinstance(data, tuple) or isinstance(data, list):
            for i, t in enumerate(data):
                if t == None:
                    self.sheet.cell(self.current_row, i+1, '')
                else:
                    self.sheet.cell(self.current_row, i+1, t)
        self.current_row += 1





if __name__ == '__main__':
    excel = ExcelReader('3011461555.xls')
    writer = ExcelWriter('test', excel.headers)
    print(writer.headers)
    for t in excel.dict_rows():
        writer.write(t)
    writer.save()
